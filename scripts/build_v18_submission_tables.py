from __future__ import annotations

"""Build the v1.8 aggregate source-data workbook and supplementary tables.

Patient-level MSK data are deliberately excluded: the upstream MSK50K license
does not permit redistribution of modified patient-level records.
"""

from pathlib import Path
import os
import shutil

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(os.environ["RCC_INPUT_BUILD_V18_SUBMISSION_TABLES_01"])
PACKAGE = ROOT / "outputs" / "evidence_audit_2026-07-16" / "submission_package_v1_8_due_diligence"
SOURCE = PACKAGE / "source_data"
TABLES = PACKAGE / "supplementary_tables"
TABLES.mkdir(exist_ok=True)

README_ROWS = [
    ("Workbook", "RCC_TKI_ICI_WES_HLA_v1_8_source_data.xlsx"),
    ("Version", "v1.8 due-diligence package, 2026-07-17"),
    ("Purpose", "Aggregate source data used to render Figures 1-5 and Supplementary Figures S1-S6."),
    ("Discovery source", "Lee et al. Mol Cancer Res 2021; PRJNA610643; official supplementary MAF/HLA tables and audited PFS."),
    ("Miao source", "cBioPortal ccrcc_dfci_2019; nivolumab-treated ccRCC mutation-only analysis."),
    ("MSK source", "cBioPortal tmb_mskcc_2018 plus msk_impact_50k_2026. Only aggregate outputs are included."),
    ("MSK restriction", "No patient-level MSK records are redistributed. Recreate from stated study IDs and analysis code, subject to source licenses."),
    ("Peptide scope", "Complete retained 63-event predicted paired universe; computational predictions, not measured immunopeptidomics."),
    ("PXD017149 scope", "Ligandomics context only; it contains no linked discovery mutations, HLA pairs, or clinical outcomes."),
    ("Not supported", "No validated biomarker, no direct HLA-A replication, no APOBEC subgroup, no global partner-allele rescue."),
]

SHEETS = [
    ("Fig1_discovery", "figure1_discovery_landscape.csv"),
    ("Fig2_sensitivity", "figure2_definition_sensitivity.csv"),
    ("Fig3_Miao_models", "figure3_miao_cox_models.csv"),
    ("Fig3_Miao_LOO", "figure3_miao_leave_one_out.csv"),
    ("Fig3_Miao_patient", "figure3_miao_patient_level.csv"),
    ("Fig4_cross_cohort", "figure4_cross_cohort_evidence_matrix.csv"),
    ("Fig4_MSK_conditional", "figure4_msk_conditional_tests.csv"),
    ("Fig4_MSK_aggregate", "figure4_msk_high_tmb_ccrcc.csv"),
    ("Fig4_MSK_interaction", "figure4_msk_interaction_tests.csv"),
    ("Fig5_complete_events", "figure5_complete_paired_event_universe.csv"),
    ("Fig5_Cterm", "figure5_cterm_enrichment.csv"),
    ("Fig5_patient", "figure5_patient_summary.csv"),
    ("S1_oncoprint", "supplementary_figure_S1_official_oncoprint.csv"),
    ("S2_HLA_audit", "supplementary_figure_S2_hla_version_audit.csv"),
    ("S3_PASS_SNV", "supplementary_figure_S3_pass_snv_counts.csv"),
    ("S3_signatures", "supplementary_figure_S3_signature_stability.csv"),
    ("S4_pair_metrics", "supplementary_figure_S4_complete_pair_metrics.csv"),
    ("S5_ligand_classes", "supplementary_figure_S5_pxd017149_classes.csv"),
    ("S5_ligand_tails", "supplementary_figure_S5_pxd017149_tails.csv"),
    ("S6_conditional", "supplementary_figure_S6_msk_conditional_tests.csv"),
    ("S6_interactions", "supplementary_figure_S6_msk_interactions.csv"),
]


def add_dataframe(workbook: Workbook, name: str, frame: pd.DataFrame) -> None:
    ws = workbook.create_sheet(name)
    ws.freeze_panes = "A2"
    for row in [list(frame.columns), *frame.fillna("").itertuples(index=False, name=None)]:
        ws.append(list(row))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 11), 45)
        for cell in column_cells[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def main() -> None:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["Field", "Value"])
    for row in README_ROWS:
        readme.append(row)
    for cell in readme[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    readme.column_dimensions["A"].width = 23
    readme.column_dimensions["B"].width = 115
    for row in readme.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for sheet, filename in SHEETS:
        add_dataframe(workbook, sheet, pd.read_csv(SOURCE / filename))

    source_book = PACKAGE / "RCC_TKI_ICI_WES_HLA_v1_8_source_data.xlsx"
    workbook.save(source_book)

    table_map = {
        "Supplementary_Table_S1_discovery_patient_level.csv": "figure1_discovery_landscape.csv",
        "Supplementary_Table_S2_discovery_definition_sensitivity.csv": "figure2_definition_sensitivity.csv",
        "Supplementary_Table_S3_Miao_models_and_LOO.csv": "figure3_miao_leave_one_out.csv",
        "Supplementary_Table_S4_MSK_aggregate_results.csv": "supplementary_figure_S6_msk_conditional_tests.csv",
        "Supplementary_Table_S5_complete_peptide_universe.csv": "figure5_complete_paired_event_universe.csv",
        "Supplementary_Table_S6_claim_evidence_matrix.csv": "../CLAIM_EVIDENCE_MATRIX.csv",
    }
    for output, source_file in table_map.items():
        path = (SOURCE / source_file) if not source_file.startswith("../") else (PACKAGE / source_file[3:])
        shutil.copyfile(path, TABLES / output)

    tables_readme = TABLES / "README.md"
    tables_readme.write_text(
        "# v1.8 supplementary tables\n\n"
        "All values are aggregate or redistributable patient-level source data. "
        "No modified patient-level MSK records are included because of the MSK50K CC BY-NC-ND 4.0 license. "
        "The Miao table is available in the source-data workbook; S3 provides the leave-one-out audit. "
        "S5 is the complete retained 63-event prediction universe, not the historically selected 29-event subset.\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
